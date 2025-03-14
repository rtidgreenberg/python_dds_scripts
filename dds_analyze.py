import argparse
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font

class Device:
    def __init__(self, name=None, ip=None):
        self.name = name
        self.ip = ip


class Participant:
    def __init__(self, name=None, key=None, device=None, path=None):
        self.name = name
        self.key = key
        self.device = device
        self.path = path


class Endpoint:
    def __init__(self, kind=None, topic_name=None, type_name=None, participant=None, reliable=None, max_sample_serialized_size=None, deadline=None, filter=None, multicast=None):
        self.kind = kind
        self.topic_name = topic_name
        self.type_name = type_name
        self.max_sample_serialized_size = max_sample_serialized_size
        self.participant = participant
        self.reliable = reliable
        self.deadline = deadline
        self.filter = filter
        self.multicast = multicast

class Domain:
    def __init__(self, domain_id=None):
        self.domain_id = domain_id
        self.participants = []
        self.endpoints = []


domains = {}
devices = {}
participants = {}
endpoints = {}


def parse_participant(participant_data):

    participant_name = None
    participant_key = None
    property_name = None
    property_value = None
    hostname = None
    filepath = None

    for child in participant_data:

        if child.tag == "key":
            participant_key = child.find("value").text

        if child.tag == "participant_name":
            # print("p_name")
            if child.find("name") is not None:
                participant_name = child.find("name").text

        if child.tag == "property":
            # print("p_name")

            for element in child.iter("element"):
                if element.find("name") is not None:
                    property_name = element.find("name").text

                if element.find("value") is not None:
                    property_value = element.find("value").text

                if property_name == "dds.sys_info.hostname":
                    hostname = property_value

                if property_name == "dds.sys_info.executable_filepath":
                    filepath = property_value

            if child.find("name") is not None:
                participant_name = child.find("name").text

        if child.tag == "default_unicast_locators":
            for element in child.iter("element"):
                if element.find("address") is not None:
                    address = element.find("address").text
                    ip_list = address.split(",")
                    last_4_ip_list = ip_list[-4:]

                if element.find("kind") is not None:
                    kind = element.find("kind").text
                    # If UDP Locator
                    if kind == "1":
                        ip_bytes = [int(hex_val, 16) for hex_val in last_4_ip_list]
                        ip_str = ".".join(map(str, ip_bytes))
                        break

    device = Device(hostname, ip_str)
    dp = Participant(participant_name, participant_key, device, filepath)

    return dp


def parse_endpoint(data_element, kind, participant):

    topic_name = None
    type_name = None
    reliable = None
    deadline = None
    content_filter = None
    multicast = None
    multicast_ip_str = None
    max_sample_serialized_size = ""
    
    for child in data_element:
        if child.tag == "topic_name":
            topic_name = child.text
            # print(topic_name)
        elif child.tag == "type_name":
            type_name = child.text
            # print(type_name)
        elif child.tag == "max_sample_serialized_size":
            max_sample_serialized_size = child.text
        elif child.tag == "reliability":
            reliable = child.find("kind").text
        elif child.tag == "deadline":
            sec = child.find("period/sec").text
            nanosec = child.find("period/nanosec").text
            if (sec == "DURATION_INFINITE_SEC" or nanosec == "DURATION_INFINITE_NSEC"):
                deadline = ""
            else:
                deadline = int(sec) + int(nanosec) / 1000000000
        elif child.tag == "content_filter_property":
            if child.find("filter_expression") is not None:
                content_filter = child.find("filter_expression").text
        elif child.tag == "multicast_locators":
            for element in child.iter("element"):
                # print("multicast")
                if element.find("address") is not None:
                    multicast = element.find("address").text
                    ip_list = multicast.split(",")
                    last_4_ip_list = ip_list[-4:]
                    ip_bytes = [int(hex_val, 16) for hex_val in last_4_ip_list]
                    multicast_ip_str = ".".join(map(str, ip_bytes))

    endpoint = Endpoint(kind, topic_name, type_name, participant, reliable, max_sample_serialized_size, deadline, content_filter, multicast_ip_str)

    return endpoint
        
        
        
def extract_tables(domain, devices_table, types_table, topics_table, 
                   topic_reliable_writers, topic_besteffort_writers, 
                   topic_reliable_readers, topic_besteffort_readers):

    endpoints = domain.endpoints
    participants = domain.participants
    
    for participant in participants:
        devices[participant.device.ip] = participant.device.name
    
    # categorize
    for endpoint in endpoints:
        if endpoint.topic_name:  # Ensure topic_name exists.
            if (
                endpoint.kind == "writer"
                and endpoint.reliable == "RELIABLE_RELIABILITY_QOS"
            ):
                topic_reliable_writers.setdefault(endpoint.topic_name, True)
            elif (
                endpoint.kind == "writer"
                and endpoint.reliable == "BEST_EFFORT_RELIABILITY_QOS"
            ):
                topic_besteffort_writers.setdefault(endpoint.topic_name, True)
            elif (
                endpoint.kind == "reader"
                and endpoint.reliable == "BEST_EFFORT_RELIABILITY_QOS"
            ):
                topic_besteffort_readers.setdefault(endpoint.topic_name, True)
            elif (
                endpoint.kind == "reader"
                and endpoint.reliable == "RELIABLE_RELIABILITY_QOS"
            ):
                topic_reliable_readers.setdefault(endpoint.topic_name, True)

        # Devices
        #    Participants
        #       Topics
        #          #datawriters
        #          #datareaders
        if not endpoint.participant.device.ip in devices_table:
            devices_table[endpoint.participant.device.ip] = [endpoint.participant.device.name, dict()]
            
        # possible that there is no participant name, if not, use key
        participant_name = endpoint.participant.name
        if participant_name == None:
            participant_name = endpoint.participant.key
        if not participant_name in devices_table[endpoint.participant.device.ip][1]:
            devices_table[endpoint.participant.device.ip][1][participant_name] = dict()
        if (
            not endpoint.topic_name
            in devices_table[endpoint.participant.device.ip][1][participant_name]
        ):
            devices_table[endpoint.participant.device.ip][1][participant_name][
                endpoint.topic_name
            ] = {"writer": 0, "reader": 0}
        devices_table[endpoint.participant.device.ip][1][participant_name][
            endpoint.topic_name
        ][endpoint.kind] += 1

        # Topic
        #   #datawriters
        #      reliable or not
        #   #datareaders
        #       reliable or not
        #       multicast (show multicast address) or not
        #       content-filters (show filter expressions) or not
        #   devices
        #       apps/processes

        if not endpoint.topic_name in topics_table:
            topics_table[endpoint.topic_name] = {
                "type": endpoint.type_name,
                "writer": {"num": 0, "reliable": "BEST_EFFORT"},
                "reader": {
                    "num": 0,
                    "reliable": "BEST_EFFORT",
                    "multicast": list(),
                    "content-filters": list(),
                },
                "devices": dict(),
                "participants": dict(),
            }
        topics_table[endpoint.topic_name][endpoint.kind]["num"] += 1
        if endpoint.reliable == "RELIABLE_RELIABILITY_QOS":
            topics_table[endpoint.topic_name][endpoint.kind]["reliable"] = "RELIABLE"
        if endpoint.kind == "reader":
            if endpoint.multicast != None:
                topics_table[endpoint.topic_name]["reader"]["multicast"].append(
                    endpoint.multicast
                )
            if endpoint.filter != None:
                topics_table[endpoint.topic_name]["reader"]["content-filters"].append(
                    endpoint.filter
                )

        if ( not endpoint.participant.device.ip in topics_table[endpoint.topic_name]["devices"]):
            topics_table[endpoint.topic_name]["devices"][endpoint.participant.device.ip] = set()

        topics_table[endpoint.topic_name]["devices"][endpoint.participant.device.ip].add(endpoint.participant.device.name)
        
        if ( not endpoint.participant.key in topics_table[endpoint.topic_name]["participants"]):
            topics_table[endpoint.topic_name]["participants"][endpoint.participant.key] = set()

        # possible that there is no participant name, so just add the participant object itself
        topics_table[endpoint.topic_name]["participants"][endpoint.participant.key].add(endpoint.participant)
        
        # types_table just has the max serialized size for each type
        if (endpoint.type_name not in types_table):
            types_table[endpoint.type_name] = 0
            
        if (endpoint.max_sample_serialized_size != ""):
            types_table[endpoint.type_name] = max(types_table[endpoint.type_name], int(endpoint.max_sample_serialized_size))

        
# Function to add a table to a worksheet
def add_table(ws, name):
    table = Table(displayName=name, ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

# Function to adjust column widths
def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Add some padding

def export_devices(wb, devices_table):
    print("\nEXPORTING DEVICES")
    header = ["ip address", "host name"]
    ws = wb.create_sheet(title="Devices")
    ws.append(header)
    for d in sorted(devices_table.keys()):
        # devices_table contain a map whose values are a list of two elements
        # the first element is the device name and the second element is a dictionary
        if isinstance(devices_table[d], list) and len(devices_table[d]) > 0:
            ws.append([d, devices_table[d][0]])
        else:
            ws.append([d, "Unknown"])
    add_table(ws, "DevicesTable")
    adjust_column_widths(ws)

def export_participants(wb, participants):
    print("\nEXPORTING PARTICIPANTS")
    header = ["host", "ip", "name", "path", "key"]
    ws = wb.create_sheet(title="Participants")
    ws.append(header)
    sorted_values = sorted(participants, key=lambda item: item.device.ip)
    for p in sorted_values:
        ws.append([p.device.name, p.device.ip, p.name, p.path, p.key])
    add_table(ws, "ParticipantsTable")
    adjust_column_widths(ws)

def export_entities(wb, endpoints, types_table, topics_table):    
    print("\nEXPORTING ENTITIES")
    header = ["topic name", "type name", "max serialized size", "kind", "host", "ip", "participant", "deadline", "reliable", "filter", "multicast"]
    ws = wb.create_sheet(title="Endpoints")
    ws.append(header)
    sorted_values = sorted(endpoints, key=lambda item: (item.topic_name, item.participant.device.ip, item.participant.key, item.kind))
    for entity in sorted_values:
        ws.append([entity.topic_name, entity.type_name, types_table[entity.type_name],
                   entity.kind, entity.participant.device.name, entity.participant.device.ip, entity.participant.name,
                   entity.deadline, entity.reliable, entity.filter, entity.multicast])
    add_table(ws, "EndpointsTable")
    adjust_column_widths(ws)

def export_topics(wb, types_table, topics_table):
    print("\nEXPORTING TOPICS")
    header = ["topic name", "max serialized size", "writers", "readers", "devices", "participants", "reliable", "filter", "multicast"]
    ws = wb.create_sheet(title="Topics")
    ws.append(header)
         
    for topic, data in sorted(topics_table.items(), key=lambda item: (-item[1]["reader"]["num"], item[0])):
        devices = len(topics_table[topic]["devices"])
        participants = len(topics_table[topic]["participants"])
        reliable = "RELIABLE" if topics_table[topic]["writer"]["reliable"] == "RELIABLE" and topics_table[topic]["reader"]["reliable"] == "RELIABLE" else "BEST_EFFORT"
        filter = len(topics_table[topic]["reader"]["content-filters"])
        ws.append([topic, types_table[topics_table[topic]["type"]], topics_table[topic]["writer"]["num"], topics_table[topic]["reader"]["num"], 
                   devices, participants, reliable, filter, 
                   f'{topics_table[topic]["reader"]["multicast"]}'])

    add_table(ws, "TopicsTable")
    adjust_column_widths(ws)

def export_analysis(wb, domain, devices_table, types_table, topics_table, topic_reliable_writers, topic_besteffort_writers,
                    topic_reliable_readers, topic_besteffort_readers):
    ws = wb.active
    ws.title = "Analysis"

    participants = domain.participants
    endpoints = domain.endpoints
    
    
    # check for Topics with writers but no readers and readers but no writers
    writers_without_readers = {}
    readers_without_writers = {}
    for endpoint1 in endpoints:
        found = False

        # possible that there is no participant name, if not, use key
        participant_name = endpoint1.participant.name
        if participant_name == None:
            participant_name = endpoint1.participant.key

        for endpoint2 in endpoints:
            if endpoint2.topic_name == endpoint1.topic_name:
                if (endpoint1.kind == "writer" and endpoint2.kind == "reader") or (
                    endpoint1.kind == "reader" and endpoint2.kind == "writer"
                ):
                    found = True
                    break
        if not found:
            if endpoint1.kind == "writer":
                if not endpoint1.topic_name in writers_without_readers:
                    writers_without_readers[endpoint1.topic_name] = list()
                writers_without_readers[endpoint1.topic_name].append(participant_name)
            if endpoint1.kind == "reader":
                if not endpoint1.topic_name in readers_without_writers:
                    readers_without_writers[endpoint1.topic_name] = list()
                readers_without_writers[endpoint1.topic_name].append(participant_name)
                
    # Check for Topics with reliable writers but only best effort readers
    topics_with_reliable_writers_but_only_best_effort_readers = []
    for topic_name in topic_reliable_writers:
        if topic_name in topic_besteffort_readers and not (
            topic_name in topic_reliable_readers
        ):
            topics_with_reliable_writers_but_only_best_effort_readers.append(topic_name)  
    
    # Check for Topics with more than DataType
    # get a list of endpoints sorted by topic name
    topics_with_multiple_types = {}
    sorted_endpoints = sorted(endpoints, key=lambda item: item.topic_name)
    
    for i in range(0, len(sorted_endpoints)-1):
        # if the next endpoint has the same topic name as the current one
        # and the type name is different, then add it to the set
        if sorted_endpoints[i].topic_name == sorted_endpoints[i+1].topic_name:
            if sorted_endpoints[i].type_name != sorted_endpoints[i+1].type_name:
                if sorted_endpoints[i].topic_name not in topics_with_multiple_types:
                    topics_with_multiple_types[sorted_endpoints[i].topic_name] = set()
                    topics_with_multiple_types[sorted_endpoints[i].topic_name].add(sorted_endpoints[i].type_name)
                topics_with_multiple_types[sorted_endpoints[i].topic_name].add(sorted_endpoints[i+1].type_name)
                
                
    ws.append([f"Devices Count: {len(devices_table)}"])
    ws.append([f"Participants Count: {len(participants)}"])
    ws.append([f"Types Count: {len(types_table)}"])
    unique_topic_names = {
        endpoint.topic_name for endpoint in endpoints if endpoint.topic_name
    }
    ws.append([f"Topics Count: {len(unique_topic_names)}"])
    count = sum(endpoint.kind == "reader" for endpoint in endpoints)
    ws.append([f"Readers Count: {count}"])
    count = sum(endpoint.kind == "writer" for endpoint in endpoints)
    ws.append([f"Writers Count: {count}"])
    ws.append([])  
    ws.append([f"DataWriters without DataReaders: {len(writers_without_readers)}"])
    ws.append([])   
    ws.append([f"DataReaders without DataWriters: {len(readers_without_writers)}"])
    ws.append([])
    ws.append([f"Topics with reliable writers but only best effort readers: {len(topics_with_reliable_writers_but_only_best_effort_readers)}"])    
    ws.append([])
    ws.append([f"Topics with more than one data type: {len(topics_with_multiple_types)}"])
    
        # Print Devices table
    ws.append([])
    ws.append([])
    ws.append(["Devices"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for device in sorted(devices_table.keys()):
        ws.append([])
        ws.append([f"IP: {device}"])
        for participant in sorted(devices_table[device][1].keys()):
            ws.append([f"    App: {participant}"])
            for topic in sorted(devices_table[device][1][participant].keys()):
                ws.append([
                    f'        {topic}', f'{devices_table[device][1][participant][topic]["writer"]} writers', f'{devices_table[device][1][participant][topic]["reader"]} readers'
                ])

    # Print Topics table
    ws.append([])
    ws.append([])
    ws.append(["Topics"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for topic in sorted(topics_table.keys()):
        ws.append([f'    {topic}',
                    f'devices: {len(topics_table[topic]["devices"])}',
                    f'participants: {len(topics_table[topic]["participants"])}',
                    f'writer: {topics_table[topic]["writer"]["num"]}',
                    f'{topics_table[topic]["writer"]["reliable"]}',
                    f'reader: {topics_table[topic]["reader"]["num"]}', 
                    f'{topics_table[topic]["reader"]["reliable"]}',
                    f'multicast - {topics_table[topic]["reader"]["multicast"]}',
                    f'content-filters - {topics_table[topic]["reader"]["content-filters"]}'
        ])


    # print out the topics with writers but no readers and readers but no writers
    ws.append([])
    ws.append([])
    ws.append([f"DataWriters without DataReaders - {len(writers_without_readers)}"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([])
    for topic_name in sorted(writers_without_readers.keys()):
        for path in sorted(writers_without_readers[topic_name]):
            ws.append([f"{topic_name} - {path}"])

    ws.append([])
    ws.append([])
    ws.append([f"DataReaders without DataWriters - {len(readers_without_writers)}"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([])
    for topic_name in sorted(readers_without_writers.keys()):
        for path in sorted(readers_without_writers[topic_name]):
            ws.append([f"{topic_name} - {path}"])

    # print out the topics with reliable writers but only best effort readers
    ws.append([])
    ws.append([])
    ws.append([f"Topics with reliable writers but only best effort readers - {len(topics_with_reliable_writers_but_only_best_effort_readers)}"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([])

    for topic in sorted(topics_with_reliable_writers_but_only_best_effort_readers):
        ws.append([topic])
    
    # print out the topics with multiple types    
    ws.append([])
    ws.append([])
    ws.append([f"Topics with more than one data type - {len(topics_with_multiple_types)}"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([])
    
    # print the topics with multiple types and the types
    for topic in sorted(topics_with_multiple_types.keys()):
        ws.append([f"{topic}"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        for type_name in topics_with_multiple_types[topic]:            
            ws.append([f"    {type_name}"])
            
        
    # Bold the headers (actually any cell with a colon in it)
    for row in ws.iter_rows():
        for cell in row:
            if ":" in str(cell.value) and not "::" in str(cell.value):
                cell.font = Font(bold=True)
    adjust_column_widths(ws)
    

def ProcessFile(filename):
    global domains

    tree = ET.parse(filename)
    root = tree.getroot()
    
    domain_participants = root.findall(".//domain_participants/value/element")

    for domain_participant in domain_participants:
        
        domain_id = domain_participant.find("domain_id").text
        
        domain = domains.get(domain_id)
        if domain is None:
            domain = Domain(domain_id)
            domains[domain_id] = domain
            
        participant_data = domain_participant.find("participant_data")
        participant = parse_participant(participant_data)
        
        domain.participants.append(participant)
        
        publications = domain_participant.findall(".//publication_data")
        for publication_data in publications:
            publication = parse_endpoint(publication_data, "writer", participant)
            domain.endpoints.append(publication)
            
        subscriptions = domain_participant.findall(".//subscription_data")
        for subscription_data in subscriptions:
            subscription = parse_endpoint(subscription_data, "reader", participant)
            domain.endpoints.append(subscription)
    
    

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process a file.")
    parser.add_argument("filename", help="Path to the file")
    args = parser.parse_args()
    filename = args.filename
   
    try:
        ProcessFile(filename)
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")

    for domain in domains.values():
        devices_table = {}
        types_table = {}
        topics_table = {}
        topic_reliable_writers = {}
        topic_besteffort_writers = {}
        topic_reliable_readers = {}
        topic_besteffort_readers = {}
        
        #process the domain data to create different tables of info
        extract_tables(domain, devices_table, types_table, topics_table, 
                       topic_reliable_writers, topic_besteffort_writers, 
                       topic_reliable_readers, topic_besteffort_readers)  
        
        # create spreadsheet and populate worksheets
        basename, _, extension = filename.rpartition(".")
        wb = openpyxl.Workbook()
            
        export_devices(wb, devices_table)
        export_participants(wb, domain.participants)
        
        export_entities(wb, domain.endpoints, types_table, topics_table)
        export_topics(wb, types_table, topics_table)
        export_analysis(wb, domain, devices_table, types_table, topics_table, topic_reliable_writers, 
                        topic_besteffort_writers, topic_reliable_readers, topic_besteffort_readers)
        
        wb.save(basename + "_domain_" + domain.domain_id + ".xlsx")
        print(f"File saved as {basename}_domain_{domain.domain_id}.xlsx")
